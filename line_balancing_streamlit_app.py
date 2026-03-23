import io
import math
from typing import Dict, List, Tuple

import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
import networkx as nx
from openpyxl import Workbook

st.markdown("""
<style>
.kpi-card {
    background: #0f172a;
    padding: 20px;
    border-radius: 12px;
    text-align: center;
    border: 1px solid #1e293b;
    box-shadow: 0 4px 10px rgba(0,0,0,0.4);
}
.kpi-title {
    color: #94a3b8;
    font-size: 14px;
}
.kpi-value {
    color: #ffffff;
    font-size: 28px;
    font-weight: bold;
}
</style>
""", unsafe_allow_html=True)

st.set_page_config(page_title="Production Line Optimization Assistant", layout="wide")
if "analysis_done" not in st.session_state:
    st.session_state.analysis_done = False

if "base_results" not in st.session_state:
    st.session_state.base_results = None

if "base_available_time" not in st.session_state:
    st.session_state.base_available_time = None

if "base_required_output" not in st.session_state:
    st.session_state.base_required_output = None

if "base_task_df" not in st.session_state:
    st.session_state.base_task_df = None

if "base_prod_df" not in st.session_state:
    st.session_state.base_prod_df = None



# -----------------------------
# Core logic
# -----------------------------
def normalize_predecessors(value) -> List[str]:
    if pd.isna(value) or str(value).strip() in {"", "-", "None", "none"}:
        return []
    return [x.strip() for x in str(value).split(",") if x.strip()]


def validate_task_data(df: pd.DataFrame) -> Tuple[bool, List[str]]:
    errors = []
    required_cols = ["Task", "Time (min)", "Immediate Predecessor"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        errors.append(f"Missing columns in TASK_DATA: {missing}")
        return False, errors

    if df["Task"].isna().any():
        errors.append("Some tasks are missing names.")

    if df["Task"].duplicated().any():
        errors.append("Duplicate task names found.")

    try:
        pd.to_numeric(df["Time (min)"], errors="raise")
    except Exception:
        errors.append("All task times must be numeric.")

    tasks = set(df["Task"].astype(str).str.strip())
    for _, row in df.iterrows():
        for pred in normalize_predecessors(row["Immediate Predecessor"]):
            if pred not in tasks:
                errors.append(f"Predecessor '{pred}' does not exist in task list.")

    # Cycle detection
    graph = {str(row["Task"]).strip(): normalize_predecessors(row["Immediate Predecessor"]) for _, row in df.iterrows()}
    visited = set()
    visiting = set()

    def dfs(node: str) -> bool:
        if node in visiting:
            return True
        if node in visited:
            return False
        visiting.add(node)
        for pred in graph[node]:
            if dfs(pred):
                return True
        visiting.remove(node)
        visited.add(node)
        return False

    for node in graph:
        if dfs(node):
            errors.append("Circular dependency detected in predecessors.")
            break

    return len(errors) == 0, list(dict.fromkeys(errors))


def extract_production_data(prod_df: pd.DataFrame) -> Tuple[float, float]:
    mapping = dict(zip(prod_df.iloc[:, 0].astype(str).str.strip(), prod_df.iloc[:, 1]))
    available_time = float(mapping["Available Time per Shift (min)"])
    required_output = float(mapping["Required Output (units/shift)"])
    return available_time, required_output


def build_task_records(df: pd.DataFrame) -> Dict[str, dict]:
    records = {}
    for _, row in df.iterrows():
        task = str(row["Task"]).strip()
        records[task] = {
            "time": float(row["Time (min)"]),
            "pred": normalize_predecessors(row["Immediate Predecessor"]),
        }
    return records


def ranked_positional_weight(tasks: Dict[str, dict]) -> Dict[str, float]:
    successors = {t: [] for t in tasks}
    for t, meta in tasks.items():
        for p in meta["pred"]:
            successors[p].append(t)

    memo = {}

    def total_successor_time(task: str) -> float:
        if task in memo:
            return memo[task]
        total = tasks[task]["time"]
        for succ in successors[task]:
            total += total_successor_time(succ)
        memo[task] = total
        return total

    return {t: total_successor_time(t) for t in tasks}


def assign_workstations(tasks: Dict[str, dict], cycle_time: float) -> List[dict]:
    weights = ranked_positional_weight(tasks)
    assigned = set()
    stations = []

    while len(assigned) < len(tasks):
        remaining = cycle_time
        station_tasks = []
        progressed = True

        while progressed:
            progressed = False
            eligible = []
            for task, meta in tasks.items():
                if task in assigned or task in station_tasks:
                    continue
                if all(pred in assigned or pred in station_tasks for pred in meta["pred"]):
                    if meta["time"] <= remaining:
                        eligible.append(task)

            if not eligible:
                break

            eligible.sort(key=lambda t: (weights[t], tasks[t]["time"], t), reverse=True)
            chosen = eligible[0]
            station_tasks.append(chosen)
            remaining -= tasks[chosen]["time"]
            progressed = True

        if not station_tasks:
            raise ValueError("No feasible assignment possible. Check cycle time and task structure.")

        for t in station_tasks:
            assigned.add(t)

        station_time = sum(tasks[t]["time"] for t in station_tasks)
        stations.append(
            {
                "Workstation": f"WS{len(stations)+1}",
                "Tasks Assigned": ", ".join(station_tasks),
                "Station Time": round(station_time, 3),
                "Idle Time": round(cycle_time - station_time, 3),
                "Utilization %": round((station_time / cycle_time) * 100, 2),
            }
        )

    return stations


def calculate_results(task_df: pd.DataFrame, available_time: float, required_output: float):
    tasks = build_task_records(task_df)
    total_work_content = sum(v["time"] for v in tasks.values())
    cycle_time = available_time / required_output
    theoretical_min_ws = math.ceil(total_work_content / cycle_time)
    stations = assign_workstations(tasks, cycle_time)
    actual_ws = len(stations)
    total_idle = round(actual_ws * cycle_time - total_work_content, 3)
    efficiency = round((total_work_content / (actual_ws * cycle_time)) * 100, 2)
    balance_delay = round(100 - efficiency, 2)

    util_values = [(s["Workstation"], s["Utilization %"]) for s in stations]
    most_loaded = max(util_values, key=lambda x: x[1])
    least_loaded = min(util_values, key=lambda x: x[1])

    insights = [
        f"{most_loaded[0]} is the highest-loaded station at {most_loaded[1]}% utilization.",
        f"{least_loaded[0]} is the most underutilized station at {least_loaded[1]}% utilization.",
    ]
    if actual_ws > theoretical_min_ws:
        insights.append("Actual workstations exceed the theoretical minimum because of precedence constraints and task packing limits.")
    else:
        insights.append("The actual number of stations matches the theoretical minimum, though idle time may still exist due to precedence structure.")

    return {
        "cycle_time": round(cycle_time, 3),
        "total_work_content": round(total_work_content, 3),
        "theoretical_min_ws": theoretical_min_ws,
        "actual_ws": actual_ws,
        "total_idle": total_idle,
        "efficiency": efficiency,
        "balance_delay": balance_delay,
        "stations": pd.DataFrame(stations),
        "insights": insights,
    }


# -----------------------------
# Visualization
# -----------------------------
def precedence_graph_figure(task_df):
    G = nx.DiGraph()

    # Build graph
    task_time_map = {}
    for _, row in task_df.iterrows():
        task = str(row["Task"]).strip()
        time_val = float(row["Time (min)"])
        preds = normalize_predecessors(row["Immediate Predecessor"])

        G.add_node(task)
        task_time_map[task] = time_val

        for pred in preds:
            G.add_edge(pred, task)

    # Assign levels using topological generations
    generations = list(nx.topological_generations(G))
    for layer, nodes in enumerate(generations):
        for node in nodes:
            G.nodes[node]["layer"] = layer

    # Layered layout: left to right
    pos = nx.multipartite_layout(G, subset_key="layer", align="horizontal")

    # Stretch layout for better readability
    pos = {node: (coord[1] * 2.8, -coord[0] * 1.8) for node, coord in pos.items()}

    # Labels with task + time
    labels = {
        node: f"{node}\n({task_time_map[node]} min)"
        for node in G.nodes()
    }

    fig, ax = plt.subplots(figsize=(12, 6))

    nx.draw_networkx_nodes(
        G,
        pos,
        node_size=3200,
        ax=ax
    )

    nx.draw_networkx_edges(
        G,
        pos,
        edge_color="black",
        arrows=True,
        arrowsize=20,
        width=1.8,
        ax=ax,
        connectionstyle="arc3,rad=0.03"
    )

    nx.draw_networkx_labels(
        G,
        pos,
        labels=labels,
        font_size=10,
        ax=ax
    )

    ax.set_title("Task Precedence Network", fontsize=16, pad=14)
    ax.axis("off")
    plt.tight_layout()
    return fig


def workstation_bar_figure(stations_df: pd.DataFrame, cycle_time: float):
    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.bar(stations_df["Workstation"], stations_df["Station Time"])
    ax.axhline(cycle_time, linestyle="--")
    ax.set_title("Workstation Load vs Cycle Time")
    ax.set_ylabel("Time (min)")
    ax.set_xlabel("Workstation")
    return fig


def export_results_excel(task_df: pd.DataFrame, prod_df: pd.DataFrame, results: dict) -> bytes:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "TASK_DATA"
    for row in [task_df.columns.tolist()] + task_df.values.tolist():
        ws1.append(row)

    ws2 = wb.create_sheet("PRODUCTION_DATA")
    for row in [prod_df.columns.tolist()] + prod_df.values.tolist():
        ws2.append(row)

    ws3 = wb.create_sheet("RESULTS_SUMMARY")
    summary_rows = [
        ["Cycle Time", results["cycle_time"]],
        ["Total Work Content", results["total_work_content"]],
        ["Theoretical Minimum Workstations", results["theoretical_min_ws"]],
        ["Actual Workstations", results["actual_ws"]],
        ["Total Idle Time", results["total_idle"]],
        ["Line Efficiency %", results["efficiency"]],
        ["Balance Delay %", results["balance_delay"]],
    ]
    for row in [["Metric", "Value"]] + summary_rows:
        ws3.append(row)

    ws4 = wb.create_sheet("WORKSTATION_ALLOCATION")
    stations_df = results["stations"]
    for row in [stations_df.columns.tolist()] + stations_df.values.tolist():
        ws4.append(row)

    ws5 = wb.create_sheet("INSIGHTS")
    ws5.append(["Engineering Insights"])
    for item in results["insights"]:
        ws5.append([item])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
    
def kpi_card(title, value):
    st.markdown(f"""
    <div class="kpi-card">
        <div class="kpi-title">{title}</div>
        <div class="kpi-value">{value}</div>
    </div>
    """, unsafe_allow_html=True)

# -----------------------------
# UI
# -----------------------------


# 🔥 ADD YOUR NEW FUNCTION HERE
def display_workstation_flow(stations_df):
    st.subheader("Workstation Flow")

    for _, row in stations_df.iterrows():
        ws = row["Workstation"]
        tasks = row["Tasks Assigned"]
        time = row["Station Time"]
        utilization = row["Utilization %"]

        # Strong color mapping
        if utilization >= 95:
            bg_color = "#b91c1c"  # strong red
        elif utilization >= 80:
            bg_color = "#ea580c"  # orange
        elif utilization >= 60:
            bg_color = "#1d4ed8"  # blue
        else:
            bg_color = "#374151"  # gray

        st.markdown(
            f"""
            <div style="
                background-color: {bg_color};
                padding: 14px;
                border-radius: 10px;
                margin-bottom: 10px;
                color: white;
                font-weight: 500;
                font-size: 15px;
            ">
            <b>{ws}</b> → [{tasks}] → {time} min  
            | Utilization: {utilization:.2f}%
            </div>
            """,
            unsafe_allow_html=True
        )
# 🔥 ADD THIS FUNCTION HERE
def apply_scenario_changes(task_df: pd.DataFrame, new_output: float = None, selected_task: str = None, new_task_time: float = None):
    scenario_task_df = task_df.copy()

    if selected_task and new_task_time is not None:
        scenario_task_df.loc[scenario_task_df["Task"] == selected_task, "Time (min)"] = float(new_task_time)

    return scenario_task_df, new_output


# -----------------------------
# UI starts here
# -----------------------------    
st.title("Production Line Optimization Assistant")
st.caption("Line balancing, workstation allocation, efficiency analysis, and export-ready results.")

with st.sidebar:
    st.header("Input Mode")
    mode = st.radio("Choose input source", ["Upload Excel", "Manual Entry"])
    st.markdown("**Expected Excel sheets**: `TASK_DATA`, `PRODUCTION_DATA`")

sample_task_df = pd.DataFrame(
    {
        "Task": ["A", "B", "C", "D", "E", "F", "G", "H", "I"],
        "Time (min)": [1.0, 2.0, 1.5, 2.5, 1.0, 2.0, 1.5, 1.0, 2.0],
        "Immediate Predecessor": ["-", "A", "A", "B", "B", "C", "D,E", "F", "G,H"],
    }
)

sample_prod_df = pd.DataFrame(
    {
        "Parameter": ["Available Time per Shift (min)", "Required Output (units/shift)"],
        "Value": [450, 100],
    }
)

if mode == "Upload Excel":
    uploaded_file = st.file_uploader("Upload line balancing Excel file", type=["xlsx"])
    if uploaded_file:
        xls = pd.ExcelFile(uploaded_file)
        if not {"TASK_DATA", "PRODUCTION_DATA"}.issubset(set(xls.sheet_names)):
            st.error("Excel file must contain TASK_DATA and PRODUCTION_DATA sheets.")
            st.stop()
        task_df = pd.read_excel(xls, sheet_name="TASK_DATA")
        prod_df = pd.read_excel(xls, sheet_name="PRODUCTION_DATA")
    else:
        task_df = sample_task_df.copy()
        prod_df = sample_prod_df.copy()
else:
    st.subheader("Manual Task Entry")
    task_df = st.data_editor(sample_task_df, num_rows="dynamic", use_container_width=True)
    st.subheader("Production Data")
    available_time = st.number_input("Available Time per Shift (min)", min_value=1.0, value=450.0, step=1.0)
    required_output = st.number_input("Required Output (units/shift)", min_value=1.0, value=100.0, step=1.0)
    prod_df = pd.DataFrame(
        {
            "Parameter": ["Available Time per Shift (min)", "Required Output (units/shift)"],
            "Value": [available_time, required_output],
        }
    )

st.divider()
col_a, col_b = st.columns([1.2, 1])

with col_a:
    st.subheader("Task Data Preview")
    st.dataframe(task_df, use_container_width=True)

with col_b:
    st.subheader("Production Data Preview")
    st.dataframe(prod_df, use_container_width=True)

run = st.button("Analyze Production Line", type="primary", use_container_width=True)

if run:
    valid, errors = validate_task_data(task_df)
    if not valid:
        for err in errors:
            st.error(err)
        st.stop()

    try:
        available_time, required_output = extract_production_data(prod_df)
    except Exception:
        st.error("Production data must contain 'Available Time per Shift (min)' and 'Required Output (units/shift)'.")
        st.stop()

    results = calculate_results(task_df, available_time, required_output)

    st.session_state.analysis_done = True
    st.session_state.base_results = results
    st.session_state.base_available_time = available_time
    st.session_state.base_required_output = required_output
    st.session_state.base_task_df = task_df.copy()
    st.session_state.base_prod_df = prod_df.copy()

if st.session_state.analysis_done:
    results = st.session_state.base_results
    available_time = st.session_state.base_available_time
    required_output = st.session_state.base_required_output
    task_df = st.session_state.base_task_df
    prod_df = st.session_state.base_prod_df

    st.success("Analysis completed successfully.")

    st.subheader("Key Performance Indicators")

    k1, k2, k3, k4 = st.columns(4)

with k1:
    kpi_card("Cycle Time", f'{results["cycle_time"]} min/unit')

with k2:
    kpi_card("Line Efficiency", f'{results["efficiency"]}%')

with k3:
    kpi_card("Actual Workstations", f'{results["actual_ws"]}')

with k4:
    kpi_card("Total Idle Time", f'{results["total_idle"]} min')


k5, k6, k7 = st.columns(3)

with k5:
    kpi_card("Total Work Content", f'{results["total_work_content"]} min')

with k6:
    kpi_card("Theoretical Min Workstations", f'{results["theoretical_min_ws"]}')

with k7:
    kpi_card("Balance Delay", f'{results["balance_delay"]}%')

    st.subheader("Workstation Allocation")
    st.dataframe(results["stations"], use_container_width=True)
    display_workstation_flow(results["stations"])

    c1, c2 = st.columns(2)
    with c1:
        st.subheader("Task Precedence Network")
        fig1 = precedence_graph_figure(task_df)
        st.pyplot(fig1)
        buf1 = io.BytesIO()
        fig1.savefig(buf1, format="png", bbox_inches="tight")
        st.download_button(
            "Download Network Diagram (PNG)",
            data=buf1.getvalue(),
            file_name="task_precedence_network.png",
            mime="image/png",
            key="download_network_png"
        )

    with c2:
        st.subheader("Workstation Load Chart")
        fig2 = workstation_bar_figure(results["stations"], results["cycle_time"])
        st.pyplot(fig2)

    

    st.subheader("Engineering Insights")


    stations_df = results["stations"]
    max_util = stations_df["Utilization %"].max()
    min_util = stations_df["Utilization %"].min()

    bottleneck_ws = stations_df[stations_df["Utilization %"] == max_util]["Workstation"].values[0]
    weak_ws = stations_df[stations_df["Utilization %"] == min_util]["Workstation"].values[0]

    st.markdown(f"""
- 🔥 **Bottleneck Station:** {bottleneck_ws} operating at {max_util}%
- ⚠️ **Underutilized Station:** {weak_ws} operating at {min_util}%
- 📉 **Idle Time Pattern:** Idle time is concentrated in lower-utilization stations
""")

    if results["actual_ws"] == results["theoretical_min_ws"]:
        st.markdown("- ✅ **Feasibility Note:** Actual stations match the theoretical minimum.")
    else:
        st.markdown("- ⚠️ **Feasibility Note:** Extra stations are required because of precedence and packing constraints.")

    st.subheader("Scenario Simulation")

    with st.form("scenario_form"):
        sc1, sc2 = st.columns(2)

        with sc1:
            scenario_output = st.number_input(
                "New Required Output (units/shift)",
                min_value=1.0,
                value=float(required_output),
                step=1.0,
                key="scenario_output"
            )

        with sc2:
            selected_task = st.selectbox(
                "Task to Modify (optional)",
                options=["None"] + task_df["Task"].astype(str).tolist(),
                key="scenario_task_select"
            )

        # Show task-time input only when a task is selected
        if selected_task != "None":
            current_task_time = float(
                task_df.loc[task_df["Task"] == selected_task, "Time (min)"].values[0]
            )

            st.info(f"Current time for Task {selected_task}: {current_task_time} min")

            scenario_task_time = st.number_input(
                f"New Time for Task {selected_task} (min)",
                min_value=0.1,
                value=current_task_time,
                step=0.1,
                key="scenario_task_time"
            )
        else:
            current_task_time = None
            scenario_task_time = None

        run_scenario = st.form_submit_button(
            "Run Scenario Simulation",
            use_container_width=True
        )
    if run_scenario:
        scenario_task_df, scenario_required_output = apply_scenario_changes(
            task_df,
            new_output=scenario_output,
            selected_task=None if selected_task == "None" else selected_task,
            new_task_time=scenario_task_time
        )

        scenario_results = calculate_results(
            scenario_task_df,
            available_time,
            scenario_required_output
        )

        st.success("Scenario analysis completed successfully.")

        st.subheader("Scenario KPI Comparison")

        s1, s2, s3, s4 = st.columns(4)

        s1.metric(
            "Cycle Time",
            f'{scenario_results["cycle_time"]} min/unit',
            delta=round(scenario_results["cycle_time"] - results["cycle_time"], 3)
        )
        s2.metric(
            "Line Efficiency",
            f'{scenario_results["efficiency"]}%',
            delta=round(scenario_results["efficiency"] - results["efficiency"], 2)
        )
        s3.metric(
            "Actual Workstations",
            scenario_results["actual_ws"],
            delta=scenario_results["actual_ws"] - results["actual_ws"]
        )
        s4.metric(
            "Total Idle Time",
            f'{scenario_results["total_idle"]} min',
            delta=round(scenario_results["total_idle"] - results["total_idle"], 3)
        )

        st.subheader("Scenario Workstation Allocation")
        st.dataframe(scenario_results["stations"], use_container_width=True)

        display_workstation_flow(scenario_results["stations"])

        st.subheader("Scenario Impact Summary")

        impact_points = []

        if scenario_required_output != required_output:
            impact_points.append(
                f"Required output changed from {required_output} to {scenario_required_output} units/shift."
            )

        if selected_task != "None":
            impact_points.append(
                f"Task {selected_task} time changed from {current_task_time} to {scenario_task_time} min."
            )

        if scenario_results["actual_ws"] > results["actual_ws"]:
            impact_points.append("Scenario requires more workstations than the base case.")
        elif scenario_results["actual_ws"] < results["actual_ws"]:
            impact_points.append("Scenario requires fewer workstations than the base case.")
        else:
            impact_points.append("Scenario uses the same number of workstations as the base case.")

        if scenario_results["efficiency"] > results["efficiency"]:
            impact_points.append("Scenario improves line efficiency.")
        elif scenario_results["efficiency"] < results["efficiency"]:
            impact_points.append("Scenario reduces line efficiency.")
        else:
            impact_points.append("Scenario keeps line efficiency unchanged.")

        for point in impact_points:
            st.markdown(f"- {point}")

    excel_bytes = export_results_excel(task_df, prod_df, results)
    st.download_button(
        "Download Results (Excel)",
        data=excel_bytes,
        file_name="line_balancing_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_excel_1"
    )
